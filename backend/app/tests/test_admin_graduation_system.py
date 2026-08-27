import io
import openpyxl
from docx import Document
from django.test import TestCase
from rest_framework.test import APIClient
from rest_framework import status

from app.models import (
    CustomUser,
    AcademicBatch,
    CourseClass,
    Student,
    Supervisor,
    SupervisorQuota,
    ProjectTopicArea,
    InternshipInfo,
    GraduationProject,
    DefenseCouncil,
    CouncilMember,
    FinalGradeSummary,
    EvaluationPolicy
)
from app.services.excel_importer import ExcelImportService
from app.services.allocation_engine import MinCostMaxFlowAllocationEngine
from app.services.reviewer_engine import ReviewerAndCouncilAllocationEngine
from app.services.document_generator import DocumentGenerationService

class AdminGraduationSystemTests(TestCase):
    def setUp(self):
        self.client = APIClient()

        # Admin user
        self.admin = CustomUser.objects.create_user(
            username="admin_utc",
            email="admin@utc.edu.vn",
            password="adminpassword123",
            user_type="admin",
            is_staff=True
        )

        # Batch
        self.batch = AcademicBatch.objects.create(
            batch_code="2026_2027_HK1",
            batch_name="Đợt ĐATN K60-K63",
            is_active=True
        )

        # Topic Area
        self.topic_software = ProjectTopicArea.objects.create(
            name="Phát triển phần mềm và ứng dụng (webApp, MobleApp)",
            code="SOFTWARE_DEV",
            is_active=True
        )

        # Supervisors & Quotas
        self.sup1_user = CustomUser.objects.create_user(
            username="gv_du",
            email="du@utc.edu.vn",
            password="password123",
            first_name="Dư",
            last_name="Nguyễn Đức",
            user_type="supervisor"
        )
        self.sup1 = Supervisor.objects.create(
            user=self.sup1_user,
            supervisor_id="GV003",
            academic_title="TS",
            department_name="CNPM"
        )
        self.quota1 = SupervisorQuota.objects.create(
            supervisor=self.sup1,
            batch=self.batch,
            viet_anh_quota=2,
            general_cntt_quota=5,
            max_total_quota=7
        )

        self.sup2_user = CustomUser.objects.create_user(
            username="gv_sao",
            email="sao@utc.edu.vn",
            password="password123",
            first_name="Sao",
            last_name="Nguyễn Kim",
            user_type="supervisor"
        )
        self.sup2 = Supervisor.objects.create(
            user=self.sup2_user,
            supervisor_id="GV008",
            academic_title="TS",
            department_name="Mạng&HTTT"
        )
        self.quota2 = SupervisorQuota.objects.create(
            supervisor=self.sup2,
            batch=self.batch,
            viet_anh_quota=2,
            general_cntt_quota=4,
            max_total_quota=6
        )

    def test_excel_importer_service(self):
        """Test importing students from Excel roster"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "IT1.659.103"

        # Headers
        ws.cell(row=6, column=1, value="Học phần: Đồ án tốt nghiệp Kỹ sư CNTT-1-1-26(N02).DA")
        ws.cell(row=9, column=1, value="STT")
        ws.cell(row=9, column=2, value="Lớp")
        ws.cell(row=9, column=3, value="Mã số SV")
        ws.cell(row=9, column=4, value="Họ đệm")
        ws.cell(row=9, column=5, value="Tên")

        # Row 1
        ws.cell(row=10, column=1, value=1)
        ws.cell(row=10, column=2, value="CNTT1 K62")
        ws.cell(row=10, column=3, value="211200001")
        ws.cell(row=10, column=4, value="Nguyễn Văn")
        ws.cell(row=10, column=5, value="An")

        # Row 2
        ws.cell(row=11, column=1, value=2)
        ws.cell(row=11, column=2, value="CNTT2 K62")
        ws.cell(row=11, column=3, value="211200002")
        ws.cell(row=11, column=4, value="Trần Thị")
        ws.cell(row=11, column=5, value="Bình")

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        res = ExcelImportService.import_students_from_excel(buffer, self.batch.id)
        self.assertTrue(res["success"])
        self.assertEqual(res["total"], 2)
        self.assertEqual(res["created"], 2)

        # Check DB
        student1 = Student.objects.get(registration_no="211200001")
        self.assertEqual(student1.user.first_name, "An")
        self.assertEqual(student1.user.last_name, "Nguyễn Văn")
        self.assertEqual(student1.course_class.class_code, "IT1.659.103")

    def test_mcmf_supervisor_allocation_engine(self):
        """Test Min-Cost Max-Flow matching engine"""
        # Create 2 students
        u1 = CustomUser.objects.create_user(username="21120001", user_type="student")
        s1 = Student.objects.create(user=u1, registration_no="21120001", academic_batch=self.batch)
        InternshipInfo.objects.create(
            student=s1,
            batch=self.batch,
            preferred_supervisor=self.sup1,
            topic_direction=self.topic_software
        )

        u2 = CustomUser.objects.create_user(username="21120002", user_type="student")
        s2 = Student.objects.create(user=u2, registration_no="21120002", academic_batch=self.batch)
        InternshipInfo.objects.create(
            student=s2,
            batch=self.batch,
            preferred_supervisor=self.sup2,
            topic_direction=self.topic_software
        )

        res = MinCostMaxFlowAllocationEngine.allocate_supervisors_for_batch(self.batch.id)
        self.assertTrue(res["success"])
        self.assertEqual(res["matched_count"], 2)
        self.assertEqual(res["unassigned_count"], 0)

        # Verify DB projects
        p1 = GraduationProject.objects.get(student=s1)
        self.assertEqual(p1.supervisor, self.sup1)
        p2 = GraduationProject.objects.get(student=s2)
        self.assertEqual(p2.supervisor, self.sup2)

    def test_reviewer_and_council_allocation_no_conflict(self):
        """Test Reviewer Allocation with strict no-conflict constraint"""
        u1 = CustomUser.objects.create_user(username="21120001", user_type="student")
        s1 = Student.objects.create(user=u1, registration_no="21120001", academic_batch=self.batch)
        p1 = GraduationProject.objects.create(
            student=s1,
            supervisor=self.sup1,  # Sup1 is GVHD
            batch=self.batch,
            topic_title_vi="ĐATN 1"
        )

        council = DefenseCouncil.objects.create(
            batch=self.batch,
            council_number=1,
            council_name="Hội đồng 1"
        )
        CouncilMember.objects.create(council=council, supervisor=self.sup2, user=self.sup2_user, role="CHAIR")

        res = ReviewerAndCouncilAllocationEngine.assign_councils_and_reviewers(self.batch.id)
        self.assertTrue(res["success"])
        self.assertEqual(res["assigned_count"], 1)

        p1.refresh_from_db()
        self.assertEqual(p1.council, council)
        self.assertEqual(p1.reviewer, self.sup2)  # Reviewer must be Sup2 (not Sup1)
        self.assertNotEqual(p1.reviewer, p1.supervisor)

    def test_document_generation_word_and_excel(self):
        """Test generating official Word Tờ trình and Excel Bảng điểm"""
        council = DefenseCouncil.objects.create(
            batch=self.batch,
            council_number=1,
            council_name="Hội đồng 1 - Kỹ sư CNTT"
        )
        CouncilMember.objects.create(council=council, supervisor=self.sup1, user=self.sup1_user, role="CHAIR")
        CouncilMember.objects.create(council=council, supervisor=self.sup2, user=self.sup2_user, role="SECRETARY")

        # Word Docx Test
        docx_buffer = DocumentGenerationService.generate_to_trinh_docx(council.id)
        self.assertGreater(len(docx_buffer.getvalue()), 1000)
        doc = Document(docx_buffer)
        text_content = " ".join([p.text for p in doc.paragraphs])
        self.assertIn("TỜ TRÌNH", text_content)
        self.assertIn("Hội đồng 1 - Kỹ sư CNTT", text_content)

        # Excel Xlsx Test
        excel_buffer = DocumentGenerationService.generate_bien_ban_excel(council.id)
        self.assertGreater(len(excel_buffer.getvalue()), 1000)
        wb = openpyxl.load_workbook(excel_buffer)
        self.assertIn("Hội đồng 1", wb.sheetnames)
