import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import io
import re
import random
import string
from django.db import transaction
from app.models import CustomUser, Student, Supervisor, SupervisorQuota, AcademicBatch, CourseClass

def generate_random_password(length=8):
    chars = string.ascii_letters + string.digits + "!@#"
    return "".join(random.choice(chars) for _ in range(length))

class ExcelImportService:
    @staticmethod
    def import_students_from_excel(
        file_obj,
        batch_id,
        password_strategy="MSSV",
        custom_fixed_password="",
        default_major="CNTT"
    ):
        """
        Import students and course classes from standard UTC portal Excel registration files
        or simple custom rosters. Supports password strategies: MSSV, FIXED, RANDOM.
        """
        try:
            batch = AcademicBatch.objects.get(id=batch_id)
        except AcademicBatch.DoesNotExist:
            return {
                "success": False,
                "error": f"AcademicBatch ID {batch_id} does not exist.",
                "total": 0,
                "created": 0,
                "updated": 0,
                "skipped": 0,
                "errors": [],
                "created_accounts": []
            }

        try:
            wb = openpyxl.load_workbook(file_obj, data_only=True)
        except Exception as ex:
            return {
                "success": False,
                "error": f"Không thể đọc file Excel: {str(ex)}",
                "total": 0,
                "created": 0,
                "updated": 0,
                "skipped": 0,
                "errors": [str(ex)],
                "created_accounts": []
            }

        total_count = 0
        created_count = 0
        updated_count = 0
        skipped_count = 0
        errors = []
        created_accounts = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if ws.max_row < 2:
                continue

            # Check if sheet contains Course Class header (e.g. Row 6: "Học phần: Đồ án tốt nghiệp-1-1-26(N41).DA")
            class_code = sheet_name.strip()
            class_name = f"Lớp {class_code}"
            class_group = ""
            program_type = "DAI_TRA"
            if default_major == "KHMT":
                program_type = "KHMT"

            for r in range(1, min(10, ws.max_row + 1)):
                val = str(ws.cell(r, 1).value or "")
                if "Học phần:" in val:
                    class_name = val.replace("Học phần:", "").strip()
                    match_grp = re.search(r'\((N\d+)\)', class_name)
                    if match_grp:
                        class_group = match_grp.group(1)
                    if "Việt - Anh" in class_name or "Việt Anh" in class_name:
                        program_type = "VIET_ANH"
                    elif "Khoa học máy tính" in class_name or "KHMT" in class_name:
                        program_type = "KHMT"
                    elif "Kỹ sư" in class_name:
                        program_type = "DAI_TRA"
                    break

            course_class, _ = CourseClass.objects.update_or_create(
                batch=batch,
                class_code=class_code,
                defaults={
                    "class_name": class_name,
                    "class_group": class_group,
                    "program_type": program_type
                }
            )

            # Find table header row
            header_row_idx = None
            col_mssv = None
            col_first_name = None
            col_last_name = None
            col_full_name = None
            col_class = None
            col_major = None
            col_phone = None
            col_email = None

            for r in range(1, min(15, ws.max_row + 1)):
                row_vals = [str(ws.cell(r, c).value or "").strip().lower() for c in range(1, ws.max_column + 1)]
                for c_idx, val in enumerate(row_vals, start=1):
                    if val in ["mã số sv", "mã sinh viên", "mssv", "mã sv", "masv"]:
                        header_row_idx = r
                        col_mssv = c_idx
                    elif val in ["họ và tên", "họ tên", "tên sinh viên", "hoten"]:
                        col_full_name = c_idx
                    elif val in ["họ và", "họ đệm", "họ"]:
                        col_last_name = c_idx
                    elif val in ["tên", "ten"]:
                        col_first_name = c_idx
                    elif val in ["lớp", "lớp sinh hoạt", "tên lớp", "lop"]:
                        col_class = c_idx
                    elif val in ["ngành", "chuyên ngành", "khoa"]:
                        col_major = c_idx
                    elif val in ["điện thoại", "sđt", "số điện thoại", "phone"]:
                        col_phone = c_idx
                    elif val in ["email", "hòm thư"]:
                        col_email = c_idx

                if header_row_idx is not None:
                    break

            if header_row_idx is None:
                continue

            # Process data rows
            for r in range(header_row_idx + 1, ws.max_row + 1):
                mssv_raw = ws.cell(r, col_mssv).value if col_mssv else None
                if not mssv_raw:
                    continue

                mssv = str(mssv_raw).strip()
                if not mssv or mssv.lower() in ["stt", "người lập", "cán bộ", "*", "tổng cộng", "tong"]:
                    continue

                # Parse name
                first_name = ""
                last_name = ""
                if col_last_name and col_first_name:
                    last_name = str(ws.cell(r, col_last_name).value or "").strip()
                    first_name = str(ws.cell(r, col_first_name).value or "").strip()
                elif col_full_name:
                    full = str(ws.cell(r, col_full_name).value or "").strip()
                    tokens = full.split()
                    if tokens:
                        first_name = tokens[-1]
                        last_name = " ".join(tokens[:-1])

                student_class_name = str(ws.cell(r, col_class).value or "").strip() if col_class else ""
                row_major = str(ws.cell(r, col_major).value or "").strip() if col_major else default_major
                phone_number = str(ws.cell(r, col_phone).value or "").strip() if col_phone else ""
                row_email = str(ws.cell(r, col_email).value or "").strip() if col_email else ""

                # Determine major and program_type
                detected_major = "KHMT" if ("khoa học máy tính" in row_major.lower() or "khmt" in row_major.lower() or "cs" in row_major.lower() or program_type == "KHMT") else "CNTT"
                if "việt - anh" in student_class_name.lower() or "việt anh" in student_class_name.lower() or "va" in student_class_name.lower():
                    row_program = "VIET_ANH"
                elif detected_major == "KHMT":
                    row_program = "KHMT"
                else:
                    row_program = "DAI_TRA"

                total_count += 1
                try:
                    with transaction.atomic():
                        username = mssv
                        email = row_email if row_email else f"{mssv.lower()}@lms.utc.edu.vn"

                        # Determine plain password
                        if password_strategy == "FIXED" and custom_fixed_password:
                            plain_password = custom_fixed_password.strip()
                        elif password_strategy == "RANDOM":
                            plain_password = generate_random_password(8)
                        else:
                            # Default MSSV
                            plain_password = mssv

                        user, user_created = CustomUser.objects.get_or_create(
                            username=username,
                            defaults={
                                "email": email,
                                "first_name": first_name,
                                "last_name": last_name,
                                "user_type": "student",
                                "is_staff": False,
                                "is_active": True
                            }
                        )

                        if user_created:
                            user.set_password(plain_password)
                            user.save()
                        else:
                            if not user.first_name and first_name:
                                user.first_name = first_name
                                user.last_name = last_name
                                user.save(update_fields=["first_name", "last_name"])

                        student, std_created = Student.objects.update_or_create(
                            user=user,
                            defaults={
                                "registration_no": mssv,
                                "department": student_class_name or course_class.class_name,
                                "course_class": course_class,
                                "academic_batch": batch,
                                "batch_no": student_class_name,
                                "phone_number": phone_number
                            }
                        )

                        action_status = "created" if (user_created or std_created) else "updated"
                        if action_status == "created":
                            created_count += 1
                        else:
                            updated_count += 1

                        created_accounts.append({
                            "id": user.id,
                            "username": username,
                            "full_name": f"{last_name} {first_name}".strip(),
                            "email": email,
                            "role": "student",
                            "major": detected_major,
                            "program_type": row_program,
                            "class_name": student_class_name or course_class.class_name,
                            "plain_password": plain_password if user_created else "(Không đổi)",
                            "status": action_status
                        })

                except Exception as ex:
                    errors.append(f"Sheet '{sheet_name}', Dòng {r}, MSSV {mssv}: {str(ex)}")
                    skipped_count += 1

        return {
            "success": True,
            "total": total_count,
            "created": created_count,
            "updated": updated_count,
            "skipped": skipped_count,
            "errors": errors,
            "created_accounts": created_accounts
        }

    @staticmethod
    def generate_student_template():
        """
        Generate Excel template for importing UTC students.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet()
        ws.title = "DS_Sinh_Vien"

        header_fill = PatternFill(start_color="0284C7", end_color="0284C7", fill_type="solid")
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        border = Border(
            left=Side(style='thin', color='D0D5DD'),
            right=Side(style='thin', color='D0D5DD'),
            top=Side(style='thin', color='D0D5DD'),
            bottom=Side(style='thin', color='D0D5DD')
        )

        headers = [
            "STT",
            "Mã sinh viên (MSSV) *",
            "Họ đệm *",
            "Tên *",
            "Ngành (CNTT/KHMT) *",
            "Lớp sinh hoạt *",
            "Số điện thoại",
            "Email"
        ]

        ws.append(headers)
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(1, col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        sample_rows = [
            [1, "201200101", "Nguyễn Văn", "An", "CNTT", "CNTT 1 - K62", "0912345671", "201200101@lms.utc.edu.vn"],
            [2, "201200102", "Trần Thị", "Bình", "CNTT", "CNTT Việt Anh - K62", "0912345672", "201200102@lms.utc.edu.vn"],
            [3, "201200103", "Lê Hoàng", "Cường", "KHMT", "KHMT 1 - K62", "0912345673", "201200103@lms.utc.edu.vn"],
            [4, "201200104", "Phạm Minh", "Đức", "CNTT", "CNTT 2 - K62", "0912345674", "201200104@lms.utc.edu.vn"],
        ]

        for row in sample_rows:
            ws.append(row)

        for row in ws.iter_rows(min_row=2, max_row=len(sample_rows) + 1, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.font = Font(name="Arial", size=10)
                cell.border = border
                if cell.column in [1, 2, 5]:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 22
        ws.column_dimensions["F"].width = 25
        ws.column_dimensions["G"].width = 18
        ws.column_dimensions["H"].width = 30

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    @staticmethod
    def export_users_to_excel(users_queryset, role="all"):
        """
        Export filtered users to Excel file.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet()
        ws.title = "Danh_Sach_Tai_Khoan"

        header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0')
        )

        headers = [
            "STT",
            "Tên đăng nhập (Username)",
            "Họ và tên",
            "Vai trò",
            "Ngành / Đơn vị",
            "Lớp / Bộ môn",
            "Khóa học / Đợt ĐATN",
            "GVHD / Đề tài",
            "Email",
            "Số điện thoại",
            "Trạng thái"
        ]

        ws.append(headers)
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(1, col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        stt = 1
        for user in users_queryset:
            full_name = user.get_full_name() or user.username
            role_display = {
                "student": "Sinh viên",
                "supervisor": "GV hướng dẫn",
                "committee_member": "Ủy viên HĐ",
                "external_examiner": "Cán bộ ngoài",
                "admin": "Quản trị viên"
            }.get(user.user_type, user.user_type)

            major_display = ""
            class_dept = ""
            batch_display = ""
            spv_topic = ""
            phone = ""

            if user.user_type == "student":
                std = getattr(user, "admin_student_profile", None)
                if std:
                    phone = std.phone_number or ""
                    class_dept = std.department or (std.course_class.class_name if std.course_class else "")
                    batch_display = std.academic_batch.batch_name if std.academic_batch else ""
                    if std.course_class and std.course_class.program_type == "KHMT":
                        major_display = "KHMT"
                    else:
                        major_display = "CNTT"

                    proj = getattr(std, "graduation_project", None)
                    if proj and proj.supervisor:
                        spv_topic = f"GVHD: {proj.supervisor.user.get_full_name()}"
            elif user.user_type == "supervisor":
                spv = getattr(user, "admin_supervisor_profile", None)
                if spv:
                    prefix = f"{spv.academic_title} " if spv.academic_title else ""
                    full_name = f"{prefix}{full_name}"
                    class_dept = spv.department_name or ""
                    phone = spv.phone_number or ""
                    major_display = "Giảng viên UTC"

            row_data = [
                stt,
                user.username,
                full_name,
                role_display,
                major_display,
                class_dept,
                batch_display,
                spv_topic,
                user.email,
                phone,
                "Hoạt động" if user.is_active else "Vô hiệu hóa"
            ]
            ws.append(row_data)
            stt += 1

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.font = Font(name="Arial", size=10)
                cell.border = border
                if cell.column in [1, 2, 4, 5, 7, 11]:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 24
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 20
        ws.column_dimensions["F"].width = 24
        ws.column_dimensions["G"].width = 22
        ws.column_dimensions["H"].width = 25
        ws.column_dimensions["I"].width = 28
        ws.column_dimensions["J"].width = 18
        ws.column_dimensions["K"].width = 16

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
